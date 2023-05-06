import ccxt
import pandas as pd
import talib
from openpyxl import Workbook

class MyExchange:
    
    def __init__(self):
        self.exchange = ccxt.binance({
            'rateLimit': 1000,
            'enableRateLimit': True,
            'futures': True
        })

    def get_ohlcv_data(self, symbol, timeframe):
        ohlcv = self.exchange.fetch_ohlcv(symbol, timeframe=timeframe)
        df = pd.DataFrame(
            ohlcv, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume'])
        df['timestamp'] = pd.to_datetime(df['timestamp'], unit='ms')
        df.set_index('timestamp', inplace=True)
        return df
    
    def get_rsi_data(self, symbol, timeframe, timeperiod=14):
        df = self.get_ohlcv_data(symbol, timeframe)
        rsi = talib.RSI(df['close'], timeperiod=timeperiod)
        df['rsi'] = rsi
        df_data = df.iloc[::-1]
        return df_data

def write_data_to_excel(data):
    wb = Workbook()
    ws = wb.active
    ws.append(['', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'])
    for i in range(0, len(data), 2):
        row_data = data[i:i+2]
        for j in range(len(row_data[0])):
            ws.cell(row=j*2+2, column=i//2+2, value=row_data[0][j])
            ws.cell(row=j*2+3, column=i//2+2, value=row_data[1][j])
    wb.save('output.xlsx')

my_exchange = MyExchange()
btc_usdt_rsi_data = my_exchange.get_rsi_data('BTC/USDT', '15m')
transpose_data = btc_usdt_rsi_data.transpose()
data_list = transpose_data.values.tolist()
write_data_to_excel(data_list)
