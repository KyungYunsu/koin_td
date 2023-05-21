import ccxt
import pandas as pd
import talib
import datetime
import os
import pytz
from openpyxl import load_workbook

# 시간 받아오기
current_time = datetime.datetime.now()

# data 디렉토리 경로
data_dir = './data'

# data 디렉토리가 존재하지 않을 경우 생성
if not os.path.exists(data_dir):
    os.makedirs(data_dir)


class MyExchange:

    def __init__(self):
        self.exchange = ccxt.binance({
            'rateLimit': 1000,
            'enableRateLimit': True,
            'options': {
                'defaultType': 'future'
            }
        })
        self.timezone = pytz.timezone('Asia/Seoul')  # 한국 시간대 설정

    def get_ohlcv_data(self, symbol, timeframe):
        ohlcv = self.exchange.fetch_ohlcv(symbol, timeframe=timeframe)
        df = pd.DataFrame(
            ohlcv, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume'])
        df['timestamp'] = pd.to_datetime(df['timestamp'], unit='ms')
        df['timestamp'] = df['timestamp'].apply(
            lambda x: pytz.utc.localize(x).astimezone(self.timezone))  # 시간대 변환
        df['timestamp'] = df['timestamp'].dt.tz_localize(None)  # 시간대 정보 제거
        df.set_index('timestamp', inplace=True)
        return df

    def get_open_close_rsi_data(self, symbol, timeframe, timeperiod=14):
        df = self.get_ohlcv_data(symbol, timeframe)
        rsi_close = talib.RSI(df['close'], timeperiod=timeperiod)
        rsi_open = talib.RSI(df['open'], timeperiod=timeperiod)
        df['close_rsi'] = rsi_close
        df['open_rsi'] = rsi_open
        df['high_profit'] = (df['high'] - df['open']) / df['open'] * 100
        df['low_profit'] = (df['low'] - df['open']) / df['open'] * 100
        df['close_profit'] = (df['close'] - df['open']) / df['open'] * 100
        df_data = df.iloc[::-1]

        return df_data


my_exchange = MyExchange()
btc_usdt_open_close_rsi_data_15m = my_exchange.get_open_close_rsi_data(
    'BTC/USDT', '15m')

btc_usdt_open_close_rsi_data_5m = my_exchange.get_open_close_rsi_data(
    'BTC/USDT', '5m')

btc_usdt_open_close_rsi_data_3m = my_exchange.get_open_close_rsi_data(
    'BTC/USDT', '3m')

# 엑셀 파일 경로 설정
file_name = 'btc_usdt_data_all.xlsx'
file_path = os.path.join(data_dir + '/' + file_name)

# 엑셀 파일이 이미 존재하는지 확인
if os.path.exists(file_path):
    # 기존 파일 열기
    book = load_workbook(file_path)
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    writer.book = book

    # 시트별로 기존 데이터 읽어오기
    existing_sheets = book.sheetnames
    existing_data = {}
    for sheet_name in existing_sheets:
        # 기존 데이터를 읽어오기 전에 해당 시트에 대한 키를 생성합니다.
        existing_data[sheet_name] = pd.DataFrame()

        existing_data[sheet_name]
        existing_data[sheet_name] = pd.read_excel(
            file_path, sheet_name=sheet_name)

    # 시트별로 새로운 데이터 추가하기
    btc_usdt_open_close_rsi_data_3m.to_excel(
        writer, sheet_name='3Minutes', index=True, float_format='%.3f', startrow=len(existing_data['3Minutes']))

    btc_usdt_open_close_rsi_data_5m.to_excel(
        writer, sheet_name='5Minutes', index=True, float_format='%.3f', startrow=len(existing_data['5Minutes']))

    btc_usdt_open_close_rsi_data_15m.to_excel(
        writer, sheet_name='15Minutes', index=True, float_format='%.3f', startrow=len(existing_data['15Minutes']))

    # 기존 데이터 추가하기
    for sheet_name, data_frame in existing_data.items():
        data_frame.to_excel(writer, sheet_name=sheet_name,
                            index=True, float_format='%.3f')

    writer.save()

else:
    # ExcelWriter 생성
    writer = pd.ExcelWriter(file_path, engine='xlsxwriter')

    # 엑셀 파일 생성
    btc_usdt_open_close_rsi_data_3m.to_excel(
        writer, sheet_name='3Minutes', index=True, float_format='%.3f')

    btc_usdt_open_close_rsi_data_5m.to_excel(
        writer, sheet_name='5Minutes', index=True, float_format='%.3f')

    btc_usdt_open_close_rsi_data_15m.to_excel(
        writer, sheet_name='15Minutes', index=True, float_format='%.3f')

    writer.save()

# 생성된 파일 경로 출력
print("엑셀 파일이 생성된 경로:", file_path)
